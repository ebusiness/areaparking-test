import os
import sys
import openpyxl as px
import time
import MySQLdb
import shutil
import re
import copy
import datetime

from urllib.parse import urljoin

from openpyxl.styles import PatternFill, colors
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import Keys

import utils


ROOT_PATH = os.getcwd()
EVIDENCE_ROOT_PATH = os.path.join(ROOT_PATH, 'evidence')
SCREEN_SHOT_NAME = 'screen_shot'
if not os.path.exists(EVIDENCE_ROOT_PATH):
    os.mkdir(EVIDENCE_ROOT_PATH)
DB_NAME = 'test_areaparking'
if sys.platform == 'linux':
    HOST_NAME = 'http://111.89.163.244:12345/'
    POS_TEST_CASE_START_ROW = 5
    POS_INPUT_START_ROW = 3
    DB_USER = 'root'
    DB_PWD = 'root'
    DB_HOST = '192.168.11.5'
else:
    HOST_NAME = 'http://127.0.0.1:8000'
    POS_TEST_CASE_START_ROW = 5
    POS_INPUT_START_ROW = 3
    DB_USER = 'root'
    DB_PWD = 'root'
    DB_HOST = 'localhost'
RED_FILL = PatternFill(start_color=colors.RED, end_color=colors.RED, fill_type='solid')
GREEN_FILL = PatternFill(start_color=colors.GREEN, end_color=colors.GREEN, fill_type='solid')


def main():
    if sys.platform == 'linux':
        options = webdriver.ChromeOptions()
        options.add_argument('headless')
        options.add_argument('no-sandbox')
        options.add_argument('disable-gpu')
        driver = webdriver.Chrome(options=options)
    else:
        driver = webdriver.Chrome(os.path.join(ROOT_PATH, 'chromedriver.exe'))

    driver.maximize_window()
    driver.get(HOST_NAME)
    try:
        driver.find_element_by_id('id_username').send_keys('admin')
        driver.find_element_by_id('id_password').send_keys('admin')
        driver.find_element_by_xpath('//*[@id="login-form"]/div[3]/button').click()
        print('adminでログインしました。')
    except NoSuchElementException:
        pass

    try:
        for path in collect_test_files():
            out_file = set_evidence_folder(path)
            test_xlsx_file(out_file, driver)
        driver.close()
    except Exception as ex:
        print(ex)
        # driver.close()
        raise ex


def collect_test_files():
    """テスト対象のファイルを収集する。

    :return:
    """
    file_list = []
    for dir_name in os.listdir(ROOT_PATH):
        path = os.path.join(ROOT_PATH, dir_name)
        if os.path.isdir(path) and dir_name.startswith('test'):
            for name in os.listdir(path):
                if name.endswith('.xlsx') and name.startswith('test'):
                    file_list.append(os.path.join(path, name))
    return file_list


def test_xlsx_file(path, driver):
    book = px.load_workbook(path)
    sheet_case = book['テストケース']
    for i in range(POS_TEST_CASE_START_ROW, sheet_case.max_row + 1):
        case_no = sheet_case['B{}'.format(i)].value
        input_db_sheet_name = sheet_case['C{}'.format(i)].value
        input_form_sheet_name = sheet_case['D{}'.format(i)].value
        expect_sheet_name = sheet_case['E{}'.format(i)].value
        result_sheet_name = sheet_case['F{}'.format(i)].value
        if input_db_sheet_name:
            input_tables(book[input_db_sheet_name])
        if input_form_sheet_name:
            input_data(book[input_form_sheet_name], driver, os.path.dirname(path))
        if expect_sheet_name:
            result_sheet = book.create_sheet(title=result_sheet_name)
            result_sheet.sheet_view.zoomScale = 85
            is_ok = expect_data(book[expect_sheet_name], driver, case_no, result_sheet)
            sheet_case['G{}'.format(i)].value = datetime.datetime.now()
            if is_ok is True:
                sheet_case['H{}'.format(i)].value = "○"
            else:
                sheet_case['H{}'.format(i)].value = "×"
                sheet_case['H{}'.format(i)].fill = RED_FILL

    book.save(path)
    book.close()


def set_evidence_folder(test_file_path):
    """エビデンスのフォルダーを設定する、あったら削除、なかったら追加する。

    :param test_file_name:
    :return:
    """
    test_file_name = os.path.basename(test_file_path)
    path = os.path.join(EVIDENCE_ROOT_PATH, os.path.splitext(test_file_name)[0])
    if os.path.exists(path):
        try:
            shutil.rmtree(path)
        except Exception as ex:
            print(path, "削除できません")
            raise ex

    # エビデンスのフォルダーを作成する。
    print("エビデンスのフォルダーを作成", path)
    time.sleep(1)
    os.mkdir(path)
    out_file_name = os.path.join(path, test_file_name)
    shutil.copy(test_file_path, out_file_name)
    return out_file_name


def input_data(sheet, driver, output_path):
    url = sheet['B1'].value
    if not url:
        return False
    driver.get(urljoin(HOST_NAME, url))
    form_name = None
    for i in range(POS_INPUT_START_ROW, sheet.max_row + 1):
        expect_kbn = sheet['A{}'.format(i)].value
        if expect_kbn == "URL:":
            url = sheet['B{}'.format(i)].value
            if not url:
                return False
            driver.get(urljoin(HOST_NAME, url))
            form_name = None
        elif expect_kbn == "FORM ID":
            form_name = sheet['B{}'.format(i)].value
        elif expect_kbn == "SEARCH":
            search_class = sheet['B{}'.format(i)].value
        elif expect_kbn == "FIELD":
            name = sheet['B{}'.format(i)].value
            value = sheet['C{}'.format(i)].value

            if form_name and name and value:
                element = driver.find_element_by_xpath('//form[@id="{}"]//*[@name="{}"]'.format(form_name, name))
                if element.tag_name == 'input':
                    input_type = element.get_attribute('type')
                    default_value = element.get_attribute('value')
                    if input_type == "checkbox":
                        label = driver.find_element_by_xpath('//form[@id="{}"]//*[@for="{}"]'.format(
                            form_name, 'id_' + name)
                        )
                        if value is True:
                            if not element.is_selected():
                                label.click()
                        elif value is False:
                            if element.is_selected():
                                label.click()
                    elif input_type == 'file':
                        element.send_keys(ROOT_PATH + value)
                    else:
                        if default_value != '':
                            element.send_keys((Keys.CONTROL , 'a'))
                        element.send_keys(value)
                elif element.tag_name == 'select':
                    data_select_id = element.get_attribute('data-select-id')
                    if data_select_id:
                        select_option_id = 'select-options-{}'.format(data_select_id)
                        # ドロップダウンリストを展開する
                        driver.find_element_by_css_selector('[data-activates={}]'.format(select_option_id)).click()
                        time.sleep(1)
                        try:
                            driver.find_element_by_css_selector('[data-activates={}]'.format(select_option_id)).click()
                        except:
                            pass
                        time.sleep(1)
                        # 指定項目を選択する。
                        xpath = '//ul[@id="{}"]//span[contains(text(), "{}")]'.format(select_option_id, value)
                        list_element = driver.find_element_by_xpath(xpath)
                        list_element.click()
                    else:
                        select_element = Select(element)
                        select_element.select_by_visible_text(value)
                elif element.tag_name == 'textarea':
                    element.send_keys((Keys.CONTROL, 'a'))
                    element.send_keys(value)
                elif element.tag_name == 'number':
                    element.send_keys(value)
            elif name and value:
                element = driver.find_element_by_xpath('//*[@id="{}"]'.format(name))
                if element.tag_name == 'input':
                    input_type = element.get_attribute('type')
                    if input_type == "checkbox":
                        label = driver.find_element_by_xpath('//*[@for="{}"]'.format(name))
                        print('label::  ', label)
                        label.click()
                    else:
                        element.send_keys((Keys.CONTROL, 'a'))
                        element.send_keys(value)
                elif element.tag_name == 'select':
                    data_select_id = element.get_attribute('data-select-id')
                    if data_select_id:
                        select_option_id = 'select-options-{}'.format(data_select_id)
                        # ドロップダウンリストを展開する
                        driver.find_element_by_css_selector('[data-activates={}]'.format(select_option_id)).click()
                        time.sleep(1)
                        try:
                            driver.find_element_by_css_selector('[data-activates={}]'.format(select_option_id)).click()
                        except:
                            pass
                        time.sleep(1)
                        # 指定項目を選択する。
                        xpath = '//ul[@id="{}"]//span[contains(text(), "{}")]'.format(select_option_id, value)
                        list_element = driver.find_element_by_xpath(xpath)
                        list_element.click()
                    else:
                        select_element = Select(element)
                        select_element.select_by_visible_text(value)
                elif element.tag_name == 'textarea':
                    element.send_keys((Keys.CONTROL, 'a'))
                    element.send_keys(value)
                elif element.tag_name == 'number':
                    element.send_keys(value)
        elif expect_kbn == "CLICK":
            try:
                xpath = sheet['B{}'.format(i)].value
                driver.find_element_by_xpath(xpath).click()
                time.sleep(1)
            except:
                pass
        elif expect_kbn == "SHOT":
            # ハードコピーを取る
            filename = sheet['B{}'.format(i)].value
            if not filename:
                filename = datetime.datetime.now().strftime('%Y%m%d%H%M%S%f')
            shot_dir = os.path.join(os.path.join(output_path, SCREEN_SHOT_NAME))
            if not os.path.exists(shot_dir):
                os.mkdir(shot_dir)
                time.sleep(1)
            index = '%04d' % len([name for name in os.listdir(shot_dir) if name.endswith('.png')])
            shot_path = os.path.join(shot_dir, "{}_{}.png".format(index, filename))
            utils.fullpage_screenshot(driver, shot_path)
        elif expect_kbn == "WORD":
            id = sheet['B{}'.format(i)].value
            value = sheet['C{}'.format(i)].value

            if search_class and id and value:
                element = driver.find_element_by_xpath('//div[@class="{}"]//*[@id="{}"]'.format(search_class, id))
                if element.tag_name == 'input':
                    input_type = element.get_attribute('type')
                    if input_type == "checkbox":
                        label = driver.find_element_by_xpath('//form[@id="{}"]//*[@for="{}"]'.format(
                            form_name, 'id_' + name)
                        )
                        if value is True:
                            if not element.is_selected():
                                label.click()
                        elif value is False:
                            if element.is_selected():
                                label.click()
                    else:
                        element.clear()
                        element.send_keys(value)
                elif element.tag_name == 'select':
                    data_select_id = element.get_attribute('data-select-id')
                    if data_select_id:
                        select_option_id = 'select-options-{}'.format(data_select_id)
                        # ドロップダウンリストを展開する
                        driver.find_element_by_css_selector('[data-activates={}]'.format(select_option_id)).click()
                        try:
                            driver.find_element_by_css_selector('[data-activates={}]'.format(select_option_id)).click()
                        except:
                            pass
                        time.sleep(1)
                        # 指定項目を選択する。
                        xpath = '//ul[@id="{}"]//span[contains(text(), "{}")]'.format(select_option_id, value)
                        list_element = driver.find_element_by_xpath(xpath)
                        list_element.click()
                        time.sleep(1)
                    else:
                        select_element = Select(element)
                        select_element.select_by_visible_text(value)
        elif expect_kbn == "ALERT":
            try:
                alt = driver.switch_to_alert()
                alt.accept()
                time.sleep(1)
            except:
                pass


def input_tables(sheet):
    for i in range(1, sheet.max_row + 1):
        expect_kbn = sheet['A{}'.format(i)].value
        if expect_kbn == "SQL":
            sql = sheet['B{}'.format(i)].value
            cnt = execute_sql(sql, None)
            print(sql, "{}件削除しました".format(cnt))
        elif expect_kbn == 'TABLE':
            table_name = sheet['B{}'.format(i)].value
            sql, col_count = get_insert_sql(sheet, i + 2, table_name)
            for r in range(i + 3, sheet.max_row + 1):
                if sheet.cell(row=r, column=2).value is None:
                    break
                vals = []
                for c in range(col_count):
                    val = sheet.cell(row=r, column=c + 2).value
                    if val == 'NULL':
                        val = None
                    vals.append(val)
                execute_sql(sql, vals)


def get_insert_sql(sheet, row_index, table_name):
    cols = []
    for c in range(2, sheet.max_column + 1):
        if sheet.cell(row=row_index, column=c).value:
            cols.append(sheet.cell(row=row_index, column=c).value)
        else:
            break
    params = ["%s"] * len(cols)
    return "INSERT INTO {0} ({1}) VALUES ({2});".format(table_name, ",".join(cols), ",".join(params)), len(cols)


def expect_data(sheet, driver, case_no, result_sheet):
    is_ok = True
    for i in range(1, sheet.max_row + 1):
        expect_kbn = sheet['A{}'.format(i)].value
        if expect_kbn == 'TABLE':
            table_name = sheet['B{}'.format(i)].value
            sql = sheet['B{}'.format(i + 1)].value
            end_row = get_expect_table_end_row(sheet, i)
            temp_flg = expect_table(sheet, table_name, sql, i, end_row, result_sheet)
            if temp_flg is False:
                is_ok = False
    return is_ok


def get_expect_table_end_row(sheet, start_row):
    row = start_row + 2
    for i in range(start_row + 2, sheet.max_row):
        if not sheet['B{}'.format(i)].value:
            break
        row += 1
    return row


def get_table_column_count(sheet, columns_index):
    cnt = 0
    for c in range(2, sheet.max_column + 1):
        if sheet.cell(row=columns_index, column=c).value:
            cnt += 1
        else:
            break
    return cnt


def expect_table(sheet, table_name, sql, start_row, end_row, result_sheet):
    """テーブルのデータを出力し、比較する。

    :param sheet:
    :param table_name: 対象のテーブル名
    :param sql:
    :param start_row: テーブル名の行インデックス
    :param end_row:
    :param result_sheet:
    :return:
    """
    is_ok = True
    if table_name and sql:
        column_count = get_table_column_count(sheet, start_row + 2)
        expect_start_row = result_sheet.max_row + 5
        expect_end_row = expect_start_row + (end_row - start_row - 3)
        result_start_row = expect_end_row + 1
        copy_and_paste_ranges(sheet, result_sheet, start_row, end_row)
        results = select_data(sql)
        # テーブルのタイトル行をコピーする 抄table的标题行
        copy_and_paste_ranges(sheet, result_sheet, start_row + 2, start_row + 2, with_data=True)
        if len(results) == 0:
            result_sheet.cell(row=result_start_row + 2, column=2).value = "0件"
        else:
            # テーブルのデータ行のフォーマットをコピーする。 copy表的data行的格式
            copy_and_paste_ranges(sheet, result_sheet, start_row + 3, start_row + 3 + len(results) - 1, skip=1, with_data=False)
            # 実行データを書き込む  写入运行的data
            for i, data in enumerate(results):
                for c, val in enumerate(data):
                    if isinstance(val, bytes):
                        val = "<bytes>"
                    elif isinstance(val, datetime.datetime):
                        val = val.strftime('%Y-%m-%d %H:%M:%S')
                    elif val is None:
                        val = "NULL"
                    else:
                        val = str(val)
                    result_sheet.cell(row=result_start_row + 2 + i, column=c + 2).value = val
        # データを比較
        for i in range(max((expect_end_row - expect_start_row + 1), len(results))):
            for c in range(column_count):
                expect_cell = result_sheet.cell(row=expect_start_row + i, column=c + 2)
                result_cell = result_sheet.cell(row=result_start_row + 2 + i, column=c + 2)
                if expect_cell.value == "9999-12-31 23:59:59" or expect_cell.value == "9999999":
                    result_cell.fill = GREEN_FILL
                elif str(expect_cell.value) != str(result_cell.value):
                    result_cell.fill = RED_FILL
                    is_ok = False
    else:
        is_ok = False

    return is_ok


def copy_and_paste_ranges(sheet, result_sheet, start_row, end_row, skip=2, with_data=True):
    dst_star_row = result_sheet.max_row + skip
    for c in range(1, sheet.max_column + 1):
        for i, r in enumerate(range(start_row, end_row + 1)):
            src_cell = sheet.cell(row=r, column=c)
            dst_cell = result_sheet.cell(row=dst_star_row + i, column=c)
            style_list = [
                copy.copy(src_cell.font),
                copy.copy(src_cell.fill),
                copy.copy(src_cell.alignment),
                src_cell.number_format,
                copy.copy(src_cell.border)
            ]
            dst_cell.font = style_list[0]
            dst_cell.fill = style_list[1]
            dst_cell.alignment = style_list[2]
            dst_cell.number_format = style_list[3]
            dst_cell.border = style_list[4]
            if with_data:
                dst_cell.value = src_cell.value


def select_data(sql):
    con = MySQLdb.connect(user=DB_USER, passwd=DB_PWD, db=DB_NAME, host=DB_HOST, charset='utf8')
    cursor = con.cursor()
    results = []
    try:
        sql = sql.rstrip().rstrip(';') + " limit 100;"
        cursor.execute(sql)
        print('SELECT:', sql)
        for row in cursor:
            results.append(row)
    except Exception as e:
        print(e)
    finally:
        cursor.close()
        con.close()
    return results


def execute_sql(sql, params):
    con = MySQLdb.connect(user=DB_USER, passwd=DB_PWD, db=DB_NAME, host=DB_HOST, charset='utf8')
    cursor = con.cursor()
    try:
        cnt = cursor.execute(sql, params)
        con.commit()
    except Exception as e:
        print("ERROR: ", sql)
        raise e
    finally:
        cursor.close()
        con.close()
    return cnt


def set_openpyxl_styles(ws, cell_range, start_row, with_border=False):
    rows = list(ws.iter_rows(cell_range))
    style_list = []
    reg = re.compile(r'\b[A-Z]{1,2}([0-9])+\b')
    dict_formulae = {}

    # we convert iterator to list for simplicity, but it's not memory efficient solution
    rows = list(rows)
    for row_index, cells in enumerate(rows):
        for col_index, cell in enumerate(cells):
            if row_index == 0:
                temp_list = [copy.copy(cell.font), copy.copy(cell.fill), copy.copy(cell.alignment), cell.number_format,]
                if with_border:
                    temp_list.append(copy.copy(cell.border))
                style_list.append(temp_list)
                # フォーミュラ
                if cell.value and cell.value[0] == '=':
                    lst = reg.findall(cell.value)
                    if lst and lst.count(lst[0]) == len(lst) and int(lst[0]) == start_row:
                        dict_formulae[col_index] = cell.value.replace(lst[0], '{0}')
            else:
                cell.font = style_list[col_index][0]
                cell.fill = style_list[col_index][1]
                cell.alignment = style_list[col_index][2]
                cell.number_format = style_list[col_index][3]
                if with_border:
                    cell.border = style_list[col_index][4]
                if cell.value and cell.value[0] == '=':
                    pass
                elif col_index in dict_formulae:
                    formulae = dict_formulae[col_index].format(start_row + row_index)
                    cell.value = formulae


if __name__ == '__main__':
    main()
